from collections import defaultdict
import bs4, requests, webbrowser
from annuncio import Annuncio
import openpyxl

totalResults=0
annunci = []
make = "ford"
model = "fiesta"
country = "I"  #Europe: D%2CA%2CB%2CE%2CF%2CI%2CL%2CN
#fuel = "B"
minPrice = 0
maxPrice = 0
minKm = 0
maxKm = 0
minYear = 0
maxYear = 0

#Create the link for the search page
def linkCreator(make, model, country, page):
    BASE_LINK = "https://www.autoscout24.it/lst/"+make+"/"+model+"?atype=C&cy="+country+"&desc=0&page="+str(page)+"&sort=standard&source=homepage_search-mask&ustate=N%2CU"
    #BASE_LINK = f"https://www.autoscout24.it/lst/{make}/{model}?atype=C&cy={country}&desc=0&page={str(page)}&sort=standard&source=homepage_search-mask&ustate=N%2CU"
    return BASE_LINK

#Take the article in the page and pass it to the list population method
def pageResearch(soup):
    #Looking for the list in the html of the page
    div_annunci=soup.find('main', class_='ListPage_main___0g2X')#Html object, html class
    article_annunci = div_annunci.find_all('article', class_='cldt-summary-full-item listing-impressions-tracking list-page-item ListItem_article__qyYw7') #Extraction of all cars of the page
    annunci = listPopulation(article_annunci)
    #printList(annunci)
    #return annunci

#Populate the list with all the announcements
def listPopulation(page):
    global annunci
    for articolo in page:
        marca = str(articolo.get('data-make'))
        modello = str(articolo.get('data-model'))
        prezzo = str(articolo.get('data-price'))
        km = str(articolo.get('data-mileage'))
        link = str((articolo.find('a', class_='ListItem_title__ndA4s ListItem_title_new_design__QIU2b Link_link__Ajn7I')).get('href'))
        anno = str(articolo.get('data-first-registration'))
        fuel = str(articolo.get('data-fuel-type'))
        region = str(articolo.get('data-listing-country'))
        type = str(articolo.get('data-vehicle-type'))
        car = Annuncio(marca, modello, prezzo, km, link, anno, fuel, region, type)
        annunci.append(car)
    return annunci

#Print the list
def printList(annunci):
    global totalResults
    totalResults += len(annunci)
    with open('lista.txt', 'a') as file:  # Open the file for appending
        for annuncio in annunci:
            file.write("----------------------------------\n")
            file.write("prezzo = " + str(annuncio.price) + "\n")
            file.write("marca = " + annuncio.make + "\n")
            file.write("modello = " + annuncio.model + "\n")
            file.write("km = " + str(annuncio.km) + "\n")
            file.write("anno = "+ str(annuncio.year) + "\n")
            file.write("paese = "+ str(annuncio.region) + "\n")
            file.write("carburante = "+ str(annuncio.fuel) + "\n")
            file.write("link = " + annuncio.link + "\n")
            file.write("score = " + str(annuncio.score) + "\n")
            file.write("----------------------------------\n")

#Calculates how many pages are avabile for the search you made
def pageCalculator(soup):
    pageNavigator = soup.find('div', class_='ListPage_pagination__4Vw9q')
    pageNumbers = pageNavigator.find_all('li', class_='pagination-item')
    number_of_page = 0
    for n in pageNumbers:
        page = int(n.get_text())
        if  page> number_of_page:
            number_of_page = page
    return number_of_page

#Make the call to the link and check if the response is correct. Returns the suop with the hrml code
def linkCall(link):
    response = requests.get(link)
    response.raise_for_status()
    soup=bs4.BeautifulSoup(response.text, 'html.parser')
    return soup

#This method split the list based on the "category" attribute
def categorization():
    global annunci
    new = []
    old =[]
    for a in annunci:
        if a.category == "new":
            new.append(a)
        else:
            old.append(a)
    minmax(old)
    for a in old:
        a.score = scoreCalculator(a.price, a.km, a.year)
    with open('lista.txt', 'w') as file:
        file.write("++++++++++++++++++++++++++++++++ NEW CARS ++++++++++++++++++++++++++++++++")
    new = sorted(new, key=lambda x: x.score, reverse=False)
    printList(new)
    with open('lista.txt', 'a') as file:
        file.write("++++++++++++++++++++++++++++++++ OLD CARS ++++++++++++++++++++++++++++++++")
    old = sorted(old, key=lambda x: x.score, reverse=False)
    #excelDataExport(old)
    printList(old)

def minmax(list):
    global minPrice
    global maxPrice
    global minKm
    global maxKm
    global minYear
    global maxYear
    minPrice = list[0].price
    maxPrice = list[0].price
    minKm = int(list[0].km)
    maxKm = int(list[0].km)
    minYear = int(list[0].year)
    maxYear = int(list[0].year)
    for a in list:
        if a.price < minPrice:
            minPrice = a.price
        if a.price > maxPrice:
            maxPrice = a.price
        if int(a.km) < minKm:
            minKm = a.km
        if int(a.km) > maxKm:
            maxKm = a.km
        if int(a.year) < minYear:
            minYear = a.year
        if int(a.year) > maxYear:
            maxYear = a.year

def scoreCalculator(price: int, km: int, year):
    global minPrice
    global maxPrice
    global minKm
    global maxKm 
    global minYear
    global maxYear
    year = int(year)
    price = int(price)
    km = int(km)
    normalizedPrice = normalizedValue(price, minPrice, maxPrice)
    normalizedKm = normalizedValue(km, minKm, maxKm)
    normalizedYear = 1 - normalizedValue(year, minYear, maxYear)
    convenience_score = (0.5 * normalizedPrice) + (0.2 * normalizedKm) + (0.3 * normalizedYear)
    return convenience_score

def normalizedValue(value, min, max):
    normalized = (value-min)/(max -min)
    return normalized




#EXCEL DATA EXPORT
def excelDataExport(lista_auto):

    # Creare un nuovo foglio di lavoro
    wb = openpyxl.Workbook()

    # Creare fogli per prezzi, chilometraggio e anni
    foglio_prezzi = wb.create_sheet(title='Prezzi')
    foglio_km = wb.create_sheet(title='Chilometraggio')
    foglio_anni = wb.create_sheet(title='Anni')

    # Scrivere l'intestazione delle colonne
    colonne_prezzi = ['prezzo']
    colonne_km = ['km']
    colonne_anni = ['anno']

    for colonna_index, colonna in enumerate(colonne_prezzi, start=1):
        foglio_prezzi.cell(row=1, column=colonna_index, value=colonna)

    for colonna_index, colonna in enumerate(colonne_km, start=1):
        foglio_km.cell(row=1, column=colonna_index, value=colonna)

    for colonna_index, colonna in enumerate(colonne_anni, start=1):
        foglio_anni.cell(row=1, column=colonna_index, value=colonna)

    # Scrivere i dati nelle celle
    for riga_index, auto in enumerate(lista_auto, start=2):
        foglio_prezzi.cell(row=riga_index, column=1, value=auto.price)
        foglio_km.cell(row=riga_index, column=1, value=auto.km)
        foglio_anni.cell(row=riga_index, column=1, value=auto.year)

    # Salva il file Excel
    wb.save('mio_file_excel.xlsx')


#Start the program
#TODO implemente the possibility tho choice wich make and model you want to search for
def start():
    global make
    global model
    global country
    link = linkCreator(make, model, country, page=1)
    soup = linkCall(link)
    pages = pageCalculator(soup)
    for n in range(pages+1):
        link = linkCreator(make, model, country, page=n)
        soup = linkCall(link)
        pageResearch(soup)
    categorization()

start()